# This script populates the /ref/comparison.xlsx file with the requested laptop comparison table.
import openpyxl
from openpyxl.styles import Font, Alignment

# Laptop data: [Model, Price, AI Rank, GPU, CPU, Memory, Storage, Display Size, Max Resolution, Ports]
laptops = [
    [
        "ASUS ROG Flow Z13 (2025)",
        "$3,299",
        4,
        "NVIDIA GeForce RTX 4070 8GB",
        "Intel Core Ultra 9 185H",
        "32GB LPDDR5X",
        "2TB PCIe 4.0 SSD",
        '13.4"',
        "2880x1800",
        "2x Thunderbolt 4, 1x USB-A 3.2, 1x HDMI 2.1, microSD, 3.5mm audio"
    ],
    [
        "ASUS ROG Flow Z13-KJP",
        "$2,499",
        5,
        "NVIDIA GeForce RTX 4060 8GB",
        "Intel Core i9-13900H",
        "16GB LPDDR5",
        "1TB PCIe 4.0 SSD",
        '13.4"',
        "3840x2400",
        "1x Thunderbolt 4, 1x USB-A 3.2, 1x HDMI 2.1, microSD, 3.5mm audio"
    ],
    [
        "ASUS ROG Strix G18 (2025)",
        "$3,999",
        3,
        "NVIDIA GeForce RTX 4080 12GB",
        "Intel Core i9-14900HX",
        "32GB DDR5",
        "2TB PCIe 4.0 SSD",
        '18.0"',
        "2560x1600",
        "2x Thunderbolt 4, 2x USB-A 3.2, 1x HDMI 2.1, 1x RJ45, 1x SD card, 3.5mm audio"
    ],
    [
        "ASUS ROG Strix Scar 18 (2025)",
        "$4,999",
        2,
        "NVIDIA GeForce RTX 4090 16GB",
        "Intel Core i9-14900HX",
        "64GB DDR5",
        "4TB PCIe 4.0 SSD",
        '18.0"',
        "2560x1600",
        "2x Thunderbolt 4, 2x USB-A 3.2, 1x HDMI 2.1, 1x RJ45, 1x SD card, 3.5mm audio"
    ],
    [
        'Apple MacBook Pro 16" M5 Max (2026)',
        "$7,199",
        1,
        "Apple M5 Max 40-core GPU",
        "Apple M5 Max 18-core CPU",
        "128GB Unified",
        "2TB SSD",
        '16.2"',
        "3456x2234",
        "3x Thunderbolt 4, 1x HDMI, 1x SDXC, 1x MagSafe 3, 1x 3.5mm audio"
    ]
]

headers = ["Model", "Price", "AI Rank (1=Best)", "GPU", "CPU", "Memory", "Storage", "Display Size", "Max Resolution", "Ports"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Laptop Comparison"

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# Write data
for row_idx, laptop in enumerate(laptops, 2):
    for col_idx, value in enumerate(laptop, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Auto-fit columns
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

wb.save("ref/comparison.xlsx")
